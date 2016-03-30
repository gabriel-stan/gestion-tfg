# -*- coding: utf-8 -*-
import re
import utils
from model.models import Comision_Evaluacion, Alumno, Tfg, Tfg_Asig, Profesor
from model.serializers import AlumnoSerializer, ProfesorSerializer, TFGSerializer
from django.contrib.auth.models import Group
from openpyxl import load_workbook
from gestion_tfgs.servicios import insert_tfg


def get_alumnos(username=None):
    try:
        if username:
            alumno = Alumno.objects.get(username=str(username))
            resul = AlumnoSerializer(alumno).data
        else:
            alumno = Alumno.objects.all()
            resul = AlumnoSerializer(alumno, many=True).data
            if len(resul) == 0:
                raise NameError("No hay alumnos almacenados")

        return dict(status=True, data=resul)
    except NameError as e:
        return dict(status=False, message=e.message)
    except Alumno.DoesNotExist:
        return dict(status=False, message="El alumno indicado no existe")


def insert_alumno(alumno):
    try:
        if not alumno.username:
            raise NameError("Error en el nombre del alumno")
        else:
            res = Alumno.objects.filter(username=alumno.username)
            if res.count() != 0:
                raise NameError("El alumno ya existe")

        if not alumno.first_name or not utils.is_string(alumno.first_name):
            raise NameError("Nombre incorrecto")

        if not alumno.last_name or not utils.is_string(alumno.last_name):
            raise NameError("Apellidos incorrectos")

        # exp reg para saber si el nick corresponde al correo de la ugr (@correo.ugr.es)
        if not re.match(r'^[a-z][_a-z0-9]+(@correo\.ugr\.es)$', alumno.username):
            raise NameError("El correo no es correcto")

        grupo_alumnos = Group.objects.get(name='Alumnos')
        alumno.save()
        grupo_alumnos.user_set.add(alumno)

        return dict(status=True, data=Alumno.objects.get(username=alumno.username))

    except NameError as e:
        return dict(status=False, message=e.message)


def update_alumno(alumno, campos):
    try:
        # comprobando username
        if 'username' in campos.keys():
            res = Alumno.objects.filter(username=campos['username'])
            if res.count() == 0:
                if not utils.is_string(campos['username']) or not re.match(r'^[a-z][_a-z0-9]+(@correo\.ugr\.es)$',
                                                                           campos['username']):
                    raise NameError("El correo no es correcto")
                else:
                    alumno.username = campos['username']
            else:
                raise NameError("El alumno indicado no existe")

        # comprobando nombre
        if 'first_name' in campos.keys():
            if campos['first_name'] == '' or not utils.is_string(campos['first_name']):
                raise NameError("Nombre incorrecto")
            else:
                alumno.first_name = campos['first_name']

        # comprobando apellidos
        if 'last_name' in campos.keys():
            if campos['last_name'] == '' or not utils.is_string(campos['last_name']):
                raise NameError("Apellidos incorrectos")
            else:
                alumno.last_name = campos['last_name']

        alumno.save()

        return dict(status=True, data=Alumno.objects.get(username=alumno.username))

    except NameError as e:
        return dict(status=False, message=e.message)
    except:
        return dict(status=False, message="El correo no es correcto")


def delete_alumno(alumno):

    try:
        Alumno.objects.get(username=alumno.username).delete()
        return dict(status=True)
    except Alumno.DoesNotExist:
        return dict(status=False, message="El alumno no existe")


def get_profesores(username=None):
    try:
        if username:
            profesor = Profesor.objects.get(username=str(username))
            resul = ProfesorSerializer(profesor).data
        else:
            profesor = Profesor.objects.all()
            resul = ProfesorSerializer(profesor, many=True).data
            if len(resul) == 0:
                raise NameError("No hay profesores almacenados")

        return dict(status=True, data=resul)
    except NameError as e:
        return dict(status=False, message=e.message)
    except Profesor.DoesNotExist:
        return dict(status=False, message="El profesor indicado no existe")


def insert_profesor(profesor):

    try:
        if not profesor.username or not (re.match(r'^[a-z][_a-z0-9]+(@ugr\.es)$', profesor.username)):
            raise NameError("El correo no es correcto")
        else:
            res = Profesor.objects.filter(username=profesor.username)
            if res.count() != 0:
                raise NameError("El profesor ya existe")

        if not profesor.first_name or not utils.is_string(profesor.first_name):
            raise NameError("Error en el nombre del profesor")

        if not profesor.last_name or not utils.is_string(profesor.last_name):
            raise NameError("Error en los apellidos del profesor")

        if not profesor.departamento or not utils.is_string(profesor.departamento):
            raise NameError("Error en el departamento")

        grupo_profesores = Group.objects.get(name='Profesores')
        profesor.save()
        grupo_profesores.user_set.add(profesor)

        return dict(status=True, data=Profesor.objects.get(username=profesor.username))

    except NameError as e:
        return dict(status=False, message=e.message)


def update_profesor(profesor, campos):
    try:
        # comprobando username
        if 'username' in campos.keys():
            res = Profesor.objects.filter(username=campos['username'])
            if res.count() == 0:
                if not (re.match(r'^[a-z][_a-z0-9]+(@ugr\.es)$', campos['username'])):
                    raise NameError("El correo no es correcto")
                else:
                    profesor.username = campos['username']
            else:
                raise NameError("No existe el profesor")

        # comprobando nombre
        if 'first_name' in campos.keys():
            if campos['first_name'] == '' or not utils.is_string(campos['first_name']):
                raise NameError("Error en el nombre del profesor")
            else:
                profesor.first_name = campos['first_name']

        # comprobando apellidos
        if 'last_name' in campos.keys():
            if campos['last_name'] == '' or not utils.is_string(campos['last_name']):
                raise NameError("Error en los apellidos del profesor")
            else:
                profesor.last_name = campos['last_name']

        # comprobando departamento
        if 'departamento' in campos.keys():
            if campos['departamento'] == '' or not utils.is_string(campos['departamento']):
                raise NameError("Error en el departamento")
            else:
                profesor.departamento = campos['departamento']

        profesor.save()

        return dict(status=True, data=Profesor.objects.get(username=profesor.username))

    except NameError as e:
        return dict(status=False, message=e.message)
    except:
        return dict(status=False, message="El correo no es correcto")


def delete_profesor(profesor):

    try:
        Profesor.objects.get(username=profesor.username).delete()
        return dict(status=True)
    except Profesor.DoesNotExist:
        return dict(status=False, message="El profesor no existe")


def formar_comision(presidente, sup_presidente, titular_1, sup_titular_1, titular_2=None, sup_titular_2=None):

    try:
        if not (utils.comprueba_profesor(presidente) and utils.comprueba_profesor(sup_presidente)) or \
                not (utils.comprueba_profesor(titular_1) and utils.comprueba_profesor(sup_titular_1)):

            raise NameError("Error en los miembros del comite")
        elif titular_2 and not (utils.comprueba_profesor(titular_2) and utils.comprueba_profesor(sup_titular_2)):
            raise NameError("Error en los suplentes del comite")
        else:
            comision = Comision_Evaluacion.objects.create(presidente=presidente, titular_1=titular_1,
                                                          titular_2=titular_2, sup_presidente=sup_presidente,
                                                          sup_titular_1=sup_titular_1, sup_titular_2=sup_titular_2)
            comision.save()
        return dict(status=True, data=comision)
    except NameError as e:
        return dict(status=False, message=e.message)


def subida_masiva(fichero, filas, p_fila, cabeceras):
    wb = load_workbook(fichero)
    ws = wb.active
    errores = []
    for i in range(p_fila, 5+filas):
        try:
            datos = dict(tipo=ws[cabeceras['tipo'] + str(i)].value,
                         titulo=ws[cabeceras['titulo'] + str(i)].value,
                         n_alumnos=ws[cabeceras['n_alumnos'] + str(i)].value,
                         descripcion=ws[cabeceras['descripcion'] + str(i)].value,
                         conocimientos_previos=ws[cabeceras['conocimientos_previos'] + str(i)].value,
                         hard_soft=ws[cabeceras['hard_soft'] + str(i)].value)
            if not datos['titulo']:
                errores.append(dict(fila=i, message='El TFG no tiene titulo'))
                continue
            datos['tutor'] = Profesor.objects.get(username=str(ws[cabeceras['tutor'] + str(i)].value))
            if ws[cabeceras['cotutor'] + str(i)].value:
                datos['cotutor'] = Profesor.objects.get(username=str(ws[cabeceras['cotutor'] + str(i)].value))
                tfg = Tfg(tipo=datos['tipo'], titulo=datos['titulo'], n_alumnos=datos['n_alumnos'],
                          descripcion=datos['descripcion'], conocimientos_previos=datos['conocimientos_previos'],
                          hard_soft=datos['hard_soft'],
                          tutor=datos['tutor'], cotutor=datos['cotutor'])
            else:
                tfg = Tfg(tipo=datos['tipo'], titulo=datos['titulo'], n_alumnos=datos['n_alumnos'],
                          descripcion=datos['descripcion'], conocimientos_previos=datos['conocimientos_previos'],
                          hard_soft=datos['hard_soft'],
                          tutor=datos['tutor'])
            insert_tfg(tfg)
        except Profesor.DoesNotExist:
            errores.append(dict(fila=i, message='El profesor no existe'))
            continue
        except Exception as e:
            errores.append(dict(fila=i, message=e.message))
            continue
    return dict(status=True, data=errores)
