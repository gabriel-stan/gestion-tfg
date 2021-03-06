__author__ = 'tonima'
from openpyxl import load_workbook
from authentication.models import Profesor, Alumno
from tfgs.models import Titulacion, Tfg_Asig, Tfg
import utils
from django.db.models import Q


class Tfgs_masivos(object):

    def __init__(self, fichero=None):
        if fichero:
            self.wb = load_workbook(fichero)
            self.ws = self.wb.active
        self.errores = []
        self.exitos = []

    def upload_file_tfg(self, u_fila, p_fila, cabeceras, titulacion):
        for i in range(p_fila, u_fila+1):
            try:
                data_tfg = self.read_data(cabeceras, i)
                self.tfg = self.check_tfg(data_tfg, i, titulacion)
                resul = Tfg.objects.simular_create_tfg(**self.tfg)
                if self.tfg is not False and resul is True:
                    self.exitos.append(dict(fila=i, tfg=self.tfg))
                else:
                    self.errores.append(dict(fila=i, message=resul))
            except Profesor.DoesNotExist:
                self.errores.append(dict(fila=i, message='El profesor no existe'))
                continue
            except Titulacion.DoesNotExist:
                self.errores.append(dict(fila=i, message='La titulacion no existe'))
                continue
            except Exception as e:
                self.errores.append(dict(fila=i, message=e.message))
                continue
        return dict(status=True, exitos=self.exitos, errores=self.errores)

    def read_data(self, cabeceras, i):
        resul = dict(tipo=self.ws[cabeceras['tipo'] + str(i)].value,
                     titulo=self.ws[cabeceras['titulo'] + str(i)].value,
                     n_alumnos=self.ws[cabeceras['n_alumnos'] + str(i)].value,
                     descripcion=self.ws[cabeceras['descripcion'] + str(i)].value,
                     conocimientos_previos=
                     self.ws[cabeceras['conocimientos_previos'] + str(i)].value,
                     hard_soft=self.ws[cabeceras['hard_soft'] + str(i)].value,
                     #titulacion=self.ws[cabeceras['titulacion'] + str(i)].value,
                     tutor=self.ws[cabeceras['tutor'] + str(i)].value,
                     cotutor=self.ws[cabeceras['cotutor'] + str(i)].value)

        return resul

    def check_tfg(self, tfg, i, titulacion):
        if not tfg.get('titulo'):
            raise Exception('El TFG no tiene titulo')
        tfg['tutor'] = Profesor.objects.get(email=tfg.get('tutor'))
        tfg['titulacion'] = Titulacion.objects.get(codigo=titulacion)
        if tfg.get('cotutor'):
            tfg['cotutor'] = Profesor.objects.get(email=str(tfg.get('cotutor')))
            tfg = dict(tipo=tfg['tipo'], titulo=tfg['titulo'], n_alumnos=tfg['n_alumnos'],
                       descripcion=tfg['descripcion'], conocimientos_previos=tfg['conocimientos_previos'],
                       hard_soft=tfg['hard_soft'],
                       tutor=tfg['tutor'].email, cotutor=tfg['cotutor'].email, titulacion=tfg['titulacion'].codigo)
        else:
            tfg = dict(tipo=tfg['tipo'], titulo=tfg['titulo'], n_alumnos=tfg['n_alumnos'],
                       descripcion=tfg['descripcion'], conocimientos_previos=tfg['conocimientos_previos'],
                       hard_soft=tfg['hard_soft'],
                       tutor=tfg['tutor'].email, titulacion=tfg['titulacion'].codigo)
        return tfg

    def upload_file_confirm(self, tfgs):
        errores = []
        for index, data_tfg in enumerate(tfgs):
            try:
                tfg = data_tfg.get('tfg')
                res = Tfg.objects.create(**tfg)
                if not res.get('status'):
                    errores.append(dict(fila=index, tfg=tfg))
            except Exception as e:
                errores.append(dict(fila=index, message=e.message))
                continue
        return dict(status=True, errores=errores)


class Tfgs_asig_masivos(Tfgs_masivos):

    def __init__(self, fichero=None):
        super(Tfgs_asig_masivos, self).__init__(fichero)

    def upload_file_tfg(self, u_fila, p_fila, cabeceras, titulacion):
        for i in range(p_fila, u_fila+1):
            try:
                data_tfg = self.read_data(cabeceras, i)
                self.tfg = self.check_tfg(data_tfg, i, titulacion)
                resul = Tfg.objects.simular_create_tfg(**self.tfg)
                if self.tfg is not False and resul is True:
                    model_tfg = Tfg(**data_tfg)
                    self.check_tfg_asig(data_tfg, cabeceras, i)
                    tfg_asig = dict(tfg=model_tfg, alumno_1=data_tfg['alumno_1'], alumno_2=data_tfg['alumno_2'],
                                    alumno_3=data_tfg['alumno_3'])
                    resul = Tfg_Asig.objects.simular_create_tfg_asig(**tfg_asig)
                    if resul is True:
                        self.exitos.append(dict(fila=i, tfg=self.tfg))
                    else:
                        self.errores.append(dict(fila=i, message=resul))
                else:
                    self.errores.append(dict(fila=i, message=resul))
            except Profesor.DoesNotExist:
                self.errores.append(dict(fila=i, message='El profesor no existe'))
                continue
            except Alumno.DoesNotExist:
                self.errores.append(dict(fila=i, message='El alumno no existe'))
                continue
            except Titulacion.DoesNotExist:
                self.errores.append(dict(fila=i, message='La titulacion no existe'))
                continue
            except Exception as e:
                self.errores.append(dict(fila=i, message=e.message))
                continue
        return dict(status=True, exitos=self.exitos, errores=self.errores)

    def check_tfg_asig(self, data_tfg, cabeceras, i):
        data_tfg['alumno_1'], self.tfg['alumno_1'] = utils.alumno_email_or_dni(
            unicode(self.ws[cabeceras['alumno_1'] + str(i)].value) if cabeceras.get('alumno_1') and \
                                                                       self.ws[cabeceras['alumno_1'] + str(i)].value \
            else None)

        self.tfg['nombre_alumno_1'] = unicode(self.ws[cabeceras['nombre_alumno_1'] + str(i)].value) if cabeceras.get('nombre_alumno_1') and \
                                                                       self.ws[cabeceras['nombre_alumno_1'] + str(i)].value \
            else None

        data_tfg['alumno_2'], self.tfg['alumno_2'] = utils.alumno_email_or_dni(
            unicode(self.ws[cabeceras['alumno_2'] + str(i)].value) if cabeceras.get('alumno_2') and \
                                                                       self.ws[cabeceras['alumno_2'] + str(i)].value \
            else None)

        self.tfg['nombre_alumno_2'] = unicode(self.ws[cabeceras['nombre_alumno_2'] + str(i)].value) if cabeceras.get('nombre_alumno_2') and \
                                                                       self.ws[cabeceras['nombre_alumno_2'] + str(i)].value \
            else None

        data_tfg['alumno_3'], self.tfg['alumno_3'] = utils.alumno_email_or_dni(
            unicode(self.ws[cabeceras['alumno_3'] + str(i)].value) if cabeceras.get('alumno_3') and \
                                                                       self.ws[cabeceras['alumno_3'] + str(i)].value \
            else None)

        self.tfg['nombre_alumno_3'] = unicode(self.ws[cabeceras['nombre_alumno_3'] + str(i)].value) if cabeceras.get('nombre_alumno_3') and \
                                                                       self.ws[cabeceras['nombre_alumno_3'] + str(i)].value \
            else None


    def upload_file_confirm(self, tfgs):
        errores = []
        for index, data_tfg in enumerate(tfgs):
            try:
                self.alumno_1 = self.get_or_create_alumno(data_tfg['tfg'].get('alumno_1'), data_tfg['tfg'].get('nombre_alumno_1')) if data_tfg['tfg'] \
                    .get('alumno_1') else None
                self.alumno_2 = self.get_or_create_alumno(data_tfg['tfg'].get('alumno_2'), data_tfg['tfg'].get('nombre_alumno_2')) if data_tfg['tfg'] \
                    .get('alumno_2') else None
                self.alumno_3 = self.get_or_create_alumno(data_tfg['tfg'].get('alumno_3'), data_tfg['tfg'].get('nombre_alumno_3')) if data_tfg['tfg'] \
                    .get('alumno_3') else None
                self.tfg = Tfg.objects.create(**data_tfg['tfg'])
                res = Tfg_Asig.objects.create(tfg=self.tfg.get('data'), alumno_1=self.alumno_1, alumno_2=self.alumno_2,
                                              alumno_3=self.alumno_3)
                if not res.get('status'):
                    errores.append(dict(fila=index, tfg=data_tfg))
            except Exception as e:
                errores.append(dict(fila=index, message=e.message))
                continue
        return dict(status=True, errores=errores)

    def get_or_create_alumno(self, alumno, nombre=None):
        if utils.is_email_alumno(alumno):
            if not Alumno.objects.filter(email=alumno if alumno else None).exists():
                Alumno.objects.create_user(email=alumno, first_name=nombre)
            try:
                return Alumno.objects.get(email=alumno)
            except Alumno.DoesNotExist:
                raise NameError('Error en el alumno %s' % alumno)
        elif utils.is_dni(alumno):
            if not Alumno.objects.filter(dni=alumno if alumno else None).exists():
                Alumno.objects.create_user(dni=alumno, first_name=nombre)
            try:
                return Alumno.objects.get(dni=alumno)
            except Alumno.DoesNotExist:
                raise NameError('Error en el alumno %s' % alumno)
        else:
            raise NameError('Error en el alumno %s' % alumno)